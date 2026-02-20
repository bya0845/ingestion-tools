from datetime import datetime, timedelta
from pathlib import Path
import langextract as lx
import openpyxl
import textwrap

from src.base_classes import InspectionEntry
from src.constants import COUNTY_MAP
from dataclasses import dataclass


@dataclass
class LangExtractor:
    # 1. Define the prompt and extraction rules
    prompt = textwrap.dedent(
        """\
        Extract all fields and provide meaningful values for each key in the following list
        [cty (count), BIN, carried, crossed
        due date, access, booked access, TOWN]. Do not make up a value if no plausible value
        is found"""
    )


LANE_CLOSED_TRIGGERS = {"WZTC", "UB60", "UB50", "UB40"}

# Zero-based column indices in Chen - August Bridges.xlsx
COL = {
    "Cty": 0,
    "BIN": 1,
    "Carried": 2,
    "Crossed": 3,
    "Due Date": 6,
    "Access": 15,
    "Booked Access": 16,
    "TOWN": 19,
}


def _format_due_date(due: str | None) -> str:
    """Converts a MMDD string to mm/dd/yy format, assuming year 2025."""
    if not isinstance(due, str):
        return ""
    due = due.strip()
    if len(due) == 4 and due.isdigit():
        return f"{due[:2]}/{due[2:]}/25"
    return ""


def _lane_closed(access: str | None) -> str:
    """Returns Y if access contains WZTC or UB60, N otherwise."""
    if not access:
        return "N"
    return "Y" if any(t in access for t in LANE_CLOSED_TRIGGERS) else "N"


def _get_week_start(date: datetime) -> datetime:
    """Returns the Sunday that starts the week containing date."""
    return date - timedelta(days=(date.weekday() + 1) % 7)


def _clean(value: object) -> str:
    """Strips whitespace and non-breaking spaces from a value."""
    if not isinstance(value, str):
        return ""
    return value.replace("\xa0", " ").strip()


def load_august_data(filepath: Path) -> list[dict]:
    """Loads all August 2025 bridge entries from the source workbook, sorted by date."""
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active
    records = []

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=20, values_only=True):
        if not any(v is not None for v in row):
            continue

        booked = row[COL["Booked Access"]]
        dates: list[datetime] = []

        if isinstance(booked, datetime) and booked.year == 2025 and booked.month == 8:
            dates = [booked]
        elif isinstance(booked, str) and "8/18" in booked and "8/19" in booked:
            dates = [datetime(2025, 8, 18), datetime(2025, 8, 19)]

        if not dates:
            continue

        cty = row[COL["Cty"]]
        access = row[COL["Access"]]
        base = {
            "team": "CHEN",
            "due_date": _format_due_date(row[COL["Due Date"]]),
            "region": 8,
            "county": COUNTY_MAP.get(cty, ""),
            "bin": row[COL["BIN"]],
            "feature_carried": _clean(row[COL["Carried"]]),
            "feature_crossed": _clean(row[COL["Crossed"]]),
            "access": _clean(access),
            "town": _clean(row[COL["TOWN"]]),
            "lane_closed": _lane_closed(access),
        }

        for d in dates:
            records.append({**base, "scheduled_date": d.strftime("%m/%d/%y"), "_date": d})

    records.sort(key=lambda r: r["_date"])
    return records


def generate_august_schedules(source_path: Path) -> list[Path]:
    """Generates weekly inspection schedule files for all August 2025 entries."""
    records = load_august_data(source_path)

    weeks: dict[datetime, list[dict]] = {}
    for record in records:
        week_sun = _get_week_start(record["_date"])
        weeks.setdefault(week_sun, []).append(record)

    output_paths = []
    for week_start, entries in sorted(weeks.items()):
        clean = [{k: v for k, v in e.items() if k != "_date"} for e in entries]
        creator = WeeklyScheduleCreator(week_start=week_start, bridge_data=clean)
        filename = f"chen_schedule_{week_start.strftime('%Y_%m_%d')}.xlsx"
        path = creator.save(filename)
        output_paths.append(path)
        print(f"Saved: {path}")

    return output_paths


if __name__ == "__main__":
    project_dir = Path(__file__).resolve().parent.parent
    source = project_dir / "data" / "Chen - August Bridges.xlsx"
    generate_august_schedules(source)
