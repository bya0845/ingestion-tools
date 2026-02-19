from dataclasses import dataclass, field
from datetime import datetime, timedelta
from enum import StrEnum
from logging import getLogger, INFO, DEBUG, basicConfig
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

basicConfig(level=DEBUG)
logger = getLogger(__name__)


class Employer(StrEnum):
    WSP_USA = "WSP USA"
    SOUTH_COL = "South Col"
    LU_ENG = "Lu Eng"


@dataclass(frozen=True, eq=False)
class ScheduleStyles:
    """Frozen class for setting schedule spreadsheet border and font styles.
    thin_border is equivalent of applying four-sided thin borders to a cell,
    these objects are initialized here for reuseability.
    """

    thin_border: Border = field(
        default_factory=lambda: Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
    )
    title_font: Font = field(
        default_factory=lambda: Font(name="Arial", size=9, bold=True)
    )
    normal_font: Font = field(default_factory=lambda: Font(name="Arial", size=9))
    table_font: Font = field(default_factory=lambda: Font(name="Arial", size=10))


@dataclass(frozen=True, eq=False)
class ScheduleDimensions:
    """These are default row and column height formattings that
    copy the Region 8 inspection schedules I used in 2025."""

    column_widths: dict[str, float] = field(
        default_factory=lambda: {
            "A": 14.1,
            "B": 18,
            "C": 18,
            "D": 13.6,
            "E": 17.1,
            "F": 16.1,
            "G": 29.7,
            "H": 34,
            "I": 25.5,
            "J": 20.6,
            "K": 15,
            "L": 13,
            "M": 13,
            "N": 13,
            "O": 13,
            "P": 13,
            "Q": 13,
            "R": 13,
            "S": 13,
        }
    )
    row_heights: dict[int | tuple[int, int], float] = field(
        default_factory=lambda: {
            (1, 22): 12.75,
            23: 4.5,
            24: 30,
            (25, 159): 15,
        }
    )
    border_rows: tuple[int, int] = (1, 22)
    border_cols: tuple[int, int] = (1, 11)


@dataclass
class BaseCreator:
    """Base class for initializing weekly schedule spreadsheet with borders and and template info."""

    title: str | None = None
    contact_info: list[dict[str, str]] = field(default_factory=list)
    inspection_teams: list[dict[str, str]] = field(default_factory=list)
    workbook: Workbook = field(default_factory=Workbook)
    worksheet: Worksheet | None = field(init=False, default=None)
    week_start: datetime | None = None
    project_dir: Path = field(init=False)
    output_dir: Path = field(init=False)
    default_filename: str = field(default="schedule.xlsx")
    styles: ScheduleStyles = field(default_factory=ScheduleStyles)
    dimensions: ScheduleDimensions = field(default_factory=ScheduleDimensions)

    def __post_init__(self):
        if not self.title:
            self.title = "Bridge Inspection Weekly Schedule"
        if not self.contact_info:
            self.contact_info = self.initialize_contacts()
        if not self.inspection_teams:
            self.inspection_teams = self.initialize_inspection_teams()
        if not self.week_start:
            self.week_start = self.get_sunday()
        self._init_directories()
        self.initialize_workbook()

    def _init_directories(self) -> None:
        try:
            self.project_dir = Path(__file__).resolve().parent.parent
        except (NameError, OSError) as e:
            logger.error(f"Failed to resolve project directory: {e}")
            self.project_dir = Path.cwd()
        self.output_dir = self.project_dir / "output"

    @staticmethod
    def get_sunday(date: datetime | None = None) -> datetime:
        if date is None:
            date = datetime.now()
        return date - timedelta(days=(date.weekday() + 1) % 7)

    def initialize_workbook(self) -> None:
        logger.info("Initializing workbook")
        ws = self.workbook.active
        if ws is None:
            raise RuntimeError("Failed to create worksheet")
        ws.title = "Region 8"
        self.worksheet = ws
        logger.debug("Worksheet created: %s", ws.title)
        self.initialize_dimensions()
        self.initialize_headings()
        self.initialize_contacts_section()
        self.initialize_teams_section()
        self.initialize_access_legend()
        self.initialize_table_headers()
        logger.info("Workbook initialization complete")

    def initialize_contacts(self) -> list[dict[str, str]]:
        return [
            {
                "name": "Salvatore Iodice",
                "role": "Project Manager",
                "office_phone": "",
                "cell_phone": "(917) 763-2519",
            },
            {
                "name": "Amy Hutcheson",
                "role": "Asst. Project Manager",
                "office_phone": "(914) 449-9038",
                "cell_phone": "917-902-0186",
            },
            {
                "name": "Karen Tomapat",
                "role": "Scheduling/Office Assistant",
                "office_phone": "(914) 449-9144",
                "cell_phone": "845-283-0224",
            },
            {
                "name": "Stephanie Santiago",
                "role": "Office Assistant",
                "office_phone": "",
                "cell_phone": "917-509-0650",
            },
            {
                "name": "Stacie Diamond",
                "role": "Asst. Project Manager",
                "office_phone": "(914) 449-9136",
                "cell_phone": "845-642-7036",
            },
            {
                "name": "Robert Seeley",
                "role": "Quality Control",
                "office_phone": "",
                "cell_phone": "(914) 262-2766",
            },
        ]

    def initialize_inspection_teams(self) -> list[dict[str, str]]:
        return [
            {
                "employer": Employer.WSP_USA,
                "team_leader": "Tom Barrell",
                "atl": "Nick Diflorio",
                "phone": "518-330-8841",
            },
            {
                "employer": Employer.WSP_USA,
                "team_leader": "Ben Kolesnik",
                "atl": "Frank Fraser",
                "phone": "845-596-7106",
            },
            {
                "employer": Employer.WSP_USA,
                "team_leader": "Oleg Shyputa",
                "atl": "Dan Rivie",
                "phone": "646-387-3354",
            },
            {
                "employer": Employer.WSP_USA,
                "team_leader": "Matt Bacon",
                "atl": "Nick Mendola",
                "phone": "774-239-9739",
            },
            {
                "employer": Employer.WSP_USA,
                "team_leader": "Kevin Milligan",
                "atl": "Christian Flores",
                "phone": "212-784-0037",
            },
            {
                "employer": Employer.WSP_USA,
                "team_leader": "Dan Hadden",
                "atl": "Dionis Demukaj",
                "phone": "845-661-6525",
            },
            {
                "employer": Employer.SOUTH_COL,
                "team_leader": "Shuangbi Chen",
                "atl": "Bo Lun Yang",
                "phone": "518-955-1990",
            },
            {
                "employer": Employer.LU_ENG,
                "team_leader": "Laura Fulford",
                "atl": "Ruzen Shafir",
                "phone": "518-577-7117",
            },
        ]

    def style_cell(
        self,
        row: int,
        col: int,
        value: str,
        font: Font,
        alignment: Alignment | None = None,
    ) -> None:
        if self.worksheet is None:
            raise RuntimeError("Worksheet not initialized")
        cell = self.worksheet.cell(row=row, column=col, value=value)
        cell.font = font
        if alignment:
            cell.alignment = alignment

    def initialize_contacts_section(self) -> None:
        logger.debug("Setting contacts section")
        if self.worksheet is None:
            raise RuntimeError("Worksheet not initialized")
        row_mappings = [(5, 6), (8, 9), (11, 12), (14, 15), (17, 18), (20, 21)]
        for i, (name_row, phone_row) in enumerate(row_mappings):
            if i >= len(self.contact_info):
                break
            contact = self.contact_info[i]
            self.style_cell(
                name_row,
                1,
                f"{contact['name']}, {contact['role']}",
                self.styles.normal_font,
            )
            if contact.get("office_phone"):
                self.style_cell(
                    phone_row,
                    1,
                    f"Office PH {contact['office_phone']}",
                    self.styles.normal_font,
                )
            if contact.get("cell_phone"):
                col = 3 if contact.get("office_phone") else 1
                label = "Cell" if contact.get("office_phone") else "Cell PH"
                alignment = Alignment(horizontal="center") if col == 3 else None
                self.style_cell(
                    phone_row,
                    col,
                    f"{label} {contact['cell_phone']}",
                    self.styles.normal_font,
                    alignment=alignment,
                )
        logger.debug("Added %d contacts", len(self.contact_info))

    def initialize_teams_section(self) -> None:
        logger.debug("Setting inspection teams section")
        if self.worksheet is None:
            raise RuntimeError("Worksheet not initialized")
        ws = self.worksheet
        ws["F3"] = "Inspection Teams:"
        ws["F3"].font = self.styles.normal_font
        start_row = 3
        for i, team in enumerate(self.inspection_teams):
            row = start_row + i
            team_info = f"{team['employer']}: {team['team_leader']}, Team Leader; {team['atl']}, ATL"
            self.style_cell(row, 7, team_info, self.styles.normal_font)
            self.style_cell(
                row,
                9,
                team["phone"],
                self.styles.normal_font,
                alignment=Alignment(horizontal="center"),
            )
        for row in range(start_row, start_row + len(self.inspection_teams)):
            ws.merge_cells(f"G{row}:H{row}")
            ws.cell(row=row, column=7).alignment = Alignment(horizontal="center")
        logger.debug("Added %d inspection teams", len(self.inspection_teams))

    def initialize_access_legend(self) -> None:
        logger.debug("Setting access legend")
        if self.worksheet is None:
            raise RuntimeError("Worksheet not initialized")
        ws = self.worksheet
        ws["G18"] = "Access Key:"
        ws["G18"].font = self.styles.normal_font
        ws["G18"].alignment = Alignment(horizontal="right")
        ws["H18"] = "W = Walking"
        ws["H18"].font = self.styles.normal_font
        ws["H19"] = "SL = Step Ladder"
        ws["H19"].font = self.styles.normal_font
        ws["H20"] = "EL = Extension Ladder"
        ws["H20"].font = self.styles.normal_font
        ws["H21"] = "BT = Bucket Truck"
        ws["H21"].font = self.styles.normal_font
        ws["H22"] = "UB = Under Bridge Unit"
        ws["H22"].font = self.styles.normal_font

    def initialize_table_headers(self) -> None:
        logger.debug("Setting table headers")
        if self.worksheet is None:
            raise RuntimeError("Worksheet not initialized")
        ws = self.worksheet
        headers = [
            "TEAM",
            "SCHEDULED DATE",
            "DUE DATE",
            "REGION",
            "COUNTY",
            "BIN",
            "FEATURE CARRIED",
            "FEATURE CROSSED",
            "ACCESS",
            "TOWN/LOC",
            "LANE CLOSED",
        ]
        header_row = 24
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = self.styles.table_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.styles.thin_border
        logger.debug("Added %d header columns", len(headers))

        for row in range(header_row + 1, header_row + 11):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = self.styles.thin_border
        logger.debug("Applied borders to rows %d-%d", header_row + 1, header_row + 10)

    def initialize_dimensions(self) -> None:
        logger.debug("Setting column widths and row heights")
        if self.worksheet is None:
            raise RuntimeError("Worksheet not initialized")
        ws = self.worksheet
        for col_letter, width in self.dimensions.column_widths.items():
            ws.column_dimensions[col_letter].width = width
        logger.debug("Set %d column widths", len(self.dimensions.column_widths))

        for key, height in self.dimensions.row_heights.items():
            if isinstance(key, tuple):
                for row in range(key[0], key[1] + 1):
                    ws.row_dimensions[row].height = height
            else:
                ws.row_dimensions[key].height = height
        logger.debug("Set row heights")

        for row in range(
            self.dimensions.border_rows[0], self.dimensions.border_rows[1] + 1
        ):
            for col in range(
                self.dimensions.border_cols[0], self.dimensions.border_cols[1] + 1
            ):
                ws.cell(row=row, column=col).border = self.styles.thin_border
        logger.debug(
            "Applied borders to rows %d-%d",
            self.dimensions.border_rows[0],
            self.dimensions.border_rows[1],
        )

    def initialize_headings(self) -> None:
        logger.debug("Setting headings section")
        if self.worksheet is None:
            raise RuntimeError("Worksheet not initialized")
        ws = self.worksheet
        ws["A1"] = "WSP USA, INC."
        ws["A1"].font = self.styles.title_font
        ws["A2"] = "NYSDOT 2025 Region 8 Bridge Inspection"
        ws["A2"].font = self.styles.title_font
        ws["A3"] = "Contract No. D037877"
        ws["A3"].font = self.styles.title_font

        ws["H1"] = "Inspection Schedule for week of:"
        ws["H1"].font = self.styles.title_font
        ws.merge_cells("H1:I1")
        ws["H1"].alignment = Alignment(horizontal="right")
        ws["J1"] = self.week_start
        ws["J1"].font = self.styles.title_font
        ws["J1"].number_format = "mm/dd/yy"
        ws["J1"].alignment = Alignment(horizontal="center")
        ws["K1"] = self.week_start + timedelta(days=6)
        ws["K1"].font = self.styles.title_font
        ws["K1"].number_format = "mm/dd/yy"
        ws["K1"].alignment = Alignment(horizontal="center")
        ws["J2"] = "(Sunday)"
        ws["J2"].font = self.styles.normal_font
        ws["J2"].alignment = Alignment(horizontal="center")
        ws["K2"] = "(Saturday)"
        ws["K2"].font = self.styles.normal_font
        ws["K2"].alignment = Alignment(horizontal="center")

        ws.merge_cells("G3:H3")

    def save(self, filename: str | None = None) -> Path:
        if filename is None:
            filename = self.default_filename
        output_path = self.output_dir / filename
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
        except PermissionError as e:
            logger.error(f"Permission denied creating directory: {e}")
            raise
        except OSError as e:
            logger.error(f"Failed to create directory: {e}")
            raise

        try:
            self.workbook.save(output_path)
            logger.info(f"Workbook saved: {output_path}")
        except PermissionError as e:
            logger.error(f"Permission denied saving file: {e}")
            raise
        except OSError as e:
            logger.error(f"Failed to save workbook: {e}")
            raise
        return output_path


def create_sample_workbook() -> Path:
    creator = BaseCreator()
    return creator.save("sample_schedule.xlsx")


if __name__ == "__main__":
    path = create_sample_workbook()
    print(f"Sample workbook created: {path}")
