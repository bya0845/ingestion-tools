from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from io import BytesIO
from logging import getLogger, DEBUG, basicConfig
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

from documents.schemas import InspectionEntry
from teams.models import Personnel, Team

basicConfig(level=DEBUG)
logger = getLogger(__name__)


@dataclass(frozen=True, eq=False)
class ScheduleStyles:
    """Frozen class for setting schedule spreadsheet border and font styles.
    thin_border is equivalent of applying four-sided thin borders to a cell,
    these objects are initialized here for reuseability.
    """

    thin_border: Border = field(
        default_factory=lambda: Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
    )
    title_font: Font = field(
        default_factory=lambda: Font(name="Arial", size=9, bold=True)
    )
    normal_font: Font = field(default_factory=lambda: Font(name="Arial", size=9))
    table_font: Font = field(default_factory=lambda: Font(name="Arial", size=10))


@dataclass(frozen=True, eq=False)
class ScheduleDimensions:
    """These are default row and column height formattings that
    copy the Region 8 inspection schedules I used in 2025."""

    column_widths: dict[str, float] = field(
        default_factory=lambda: {
            "A": 14.1,
            "B": 18,
            "C": 18,
            "D": 13.6,
            "E": 17.1,
            "F": 16.1,
            "G": 29.7,
            "H": 34,
            "I": 25.5,
            "J": 20.6,
            "K": 15,
            "L": 13,
            "M": 13,
            "N": 13,
            "O": 13,
            "P": 13,
            "Q": 13,
            "R": 13,
            "S": 13,
        }
    )
    row_heights: dict[int | tuple[int, int], float] = field(
        default_factory=lambda: {
            (1, 22): 12.75,
            23: 4.5,
            24: 30,
            (25, 159): 15,
        }
    )
    border_rows: tuple[int, int] = (1, 22)
    border_cols: tuple[int, int] = (1, 11)


@dataclass
class BaseCreator:
    """Base class for initializing weekly schedule spreadsheet with borders and template info."""

    title: str | None = None
    contact_info: tuple[Personnel, ...] = field(
        default_factory=lambda: tuple(Personnel.objects.all())
    )
    inspection_teams: list[Team] = field(default_factory=list)
    workbook: Workbook = field(default_factory=Workbook)
    worksheet: Worksheet | None = field(init=False, default=None)
    week_start: datetime | None = None
    project_dir: Path = field(init=False)
    output_dir: Path = field(init=False)
    team_name: str = "Chen"
    output_dir_override: Path | None = None
    region: str = "8"
    sheet_title: str = f"Region {region}"
    default_filename: str = field(init=False)
    styles: ScheduleStyles = field(default_factory=ScheduleStyles)
    dimensions: ScheduleDimensions = field(default_factory=ScheduleDimensions)

    def __post_init__(self):
        if not self.title:
            self.title = "Bridge Inspection Weekly Schedule"
        if not self.inspection_teams:
            self.inspection_teams = self.initialize_inspection_teams()
        if not self.week_start:
            self.week_start = self.get_sunday()
        week_str = self.week_start.strftime("%-m-%-d-%y")
        self.default_filename = (
            f"{self.team_name} Region {self.region} Bridge Inspection"
            f" Weekly Schedule - Week of {week_str}.xlsx"
        )
        self._init_directories()
        self.initialize_workbook()

    def _init_directories(self) -> None:
        try:
            self.project_dir = Path(__file__).resolve().parent.parent
        except (NameError, OSError) as e:
            logger.error(f"Failed to resolve project directory: {e}")
            self.project_dir = Path.cwd()
        self.output_dir = self.output_dir_override if self.output_dir_override else self.project_dir / "output"

    @staticmethod
    def get_sunday(date: datetime | None = None) -> datetime:
        """Returns the Sunday that starts the week containing date."""
        if date is None:
            date = datetime.now()
        return date - timedelta(days=(date.weekday() + 1) % 7)

    def initialize_workbook(self) -> None:
        """Initializes the workbook worksheet, dimensions, and all header sections."""
        logger.info("Initializing workbook")
        ws = self.workbook.active
        if ws is None:
            raise RuntimeError("Failed to create worksheet")
        ws.title = self.sheet_title
        self.worksheet = ws
        logger.debug("Worksheet created: %s", ws.title)
        self.initialize_dimensions()
        self.initialize_headings()
        self.initialize_contacts_section()
        self.initialize_teams_section()
        self.initialize_access_legend()
        self.initialize_table_headers()
        logger.info("Workbook initialization complete")

    def initialize_inspection_teams(self) -> list[Team]:
        """Loads all inspection teams from the database."""
        return list(Team.objects.all())

    def style_cell(
        self,
        row: int,
        col: int,
        value: str,
        font: Font,
        alignment: Alignment | None = None,
    ) -> None:
        """Writes a value to a cell and applies font and optional alignment."""
        if self.worksheet is None:
            raise RuntimeError("Worksheet not initialized")
        cell = self.worksheet.cell(row=row, column=col, value=value)
        cell.font = font
        if alignment:
            cell.alignment = alignment

    def initialize_contacts_section(self) -> None:
        """Populates the contacts section with personnel names and phone numbers."""
        logger.debug("Setting contacts section")
        if self.worksheet is None:
            raise RuntimeError("Worksheet not initialized")
        row_mappings = [(5, 6), (8, 9), (11, 12), (14, 15), (17, 18), (20, 21)]
        for i, (name_row, phone_row) in enumerate(row_mappings):
            if i >= len(self.contact_info):
                break
            contact = self.contact_info[i]
            self.style_cell(
                name_row,
                1,
                f"{contact.name}, {contact.role}",
                self.styles.normal_font,
            )
            if contact.office_phone:
                self.style_cell(
                    phone_row,
                    1,
                    f"Office PH {contact.office_phone}",
                    self.styles.normal_font,
                )
            if contact.cell_phone:
                col = 3 if contact.office_phone else 1
                label = "Cell" if contact.office_phone else "Cell PH"
                alignment = Alignment(horizontal="center") if col == 3 else None
                self.style_cell(
                    phone_row,
                    col,
                    f"{label} {contact.cell_phone}",
                    self.styles.normal_font,
                    alignment=alignment,
                )
        logger.debug("Added %d contacts", len(self.contact_info))

    def initialize_teams_section(self) -> None:
        """Populates the inspection teams section with team names and phone numbers."""
        logger.debug("Setting inspection teams section")
        if self.worksheet is None:
            raise RuntimeError("Worksheet not initialized")
        ws = self.worksheet
        ws["F3"] = "Inspection Teams:"
        ws["F3"].font = self.styles.normal_font
        start_row = 3
        for i, team in enumerate(self.inspection_teams):
            row = start_row + i
            self.style_cell(row, 7, str(team), self.styles.normal_font)
            self.style_cell(
                row,
                9,
                team.phone,
                self.styles.normal_font,
                alignment=Alignment(horizontal="center"),
            )
        for row in range(start_row, start_row + len(self.inspection_teams)):
            ws.merge_cells(f"G{row}:H{row}")
            ws.cell(row=row, column=7).alignment = Alignment(horizontal="center")
        logger.debug("Added %d inspection teams", len(self.inspection_teams))

    def initialize_access_legend(self) -> None:
        """Writes the access method key to the worksheet."""
        logger.debug("Setting access legend")
        if self.worksheet is None:
            raise RuntimeError("Worksheet not initialized")
        ws = self.worksheet
        ws["G18"] = "Access Key:"
        ws["G18"].font = self.styles.normal_font
        ws["G18"].alignment = Alignment(horizontal="right")
        ws["H18"] = "W = Walking"
        ws["H18"].font = self.styles.normal_font
        ws["H19"] = "SL = Step Ladder"
        ws["H19"].font = self.styles.normal_font
        ws["H20"] = "EL = Extension Ladder"
        ws["H20"].font = self.styles.normal_font
        ws["H21"] = "BT = Bucket Truck"
        ws["H21"].font = self.styles.normal_font
        ws["H22"] = "UB = Under Bridge Unit"
        ws["H22"].font = self.styles.normal_font

    def initialize_table_headers(self) -> None:
        """Writes the data table header row and applies borders to the data rows below it."""
        logger.debug("Setting table headers")
        if self.worksheet is None:
            raise RuntimeError("Worksheet not initialized")
        ws = self.worksheet
        headers = [
            "TEAM",
            "SCHEDULED DATE",
            "DUE DATE",
            "REGION",
            "COUNTY",
            "BIN",
            "FEATURE CARRIED",
            "FEATURE CROSSED",
            "ACCESS",
            "TOWN/LOC",
            "LANE CLOSED",
        ]
        header_row = 24
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = self.styles.table_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.styles.thin_border
        logger.debug("Added %d header columns", len(headers))

        for row in range(header_row + 1, header_row + 11):
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = self.styles.thin_border
        logger.debug("Applied borders to rows %d-%d", header_row + 1, header_row + 10)

    def initialize_dimensions(self) -> None:
        """Applies column widths, row heights, and border grid to the worksheet."""
        logger.debug("Setting column widths and row heights")
        if self.worksheet is None:
            raise RuntimeError("Worksheet not initialized")
        ws = self.worksheet
        for col_letter, width in self.dimensions.column_widths.items():
            ws.column_dimensions[col_letter].width = width
        logger.debug("Set %d column widths", len(self.dimensions.column_widths))

        for key, height in self.dimensions.row_heights.items():
            if isinstance(key, tuple):
                for row in range(key[0], key[1] + 1):
                    ws.row_dimensions[row].height = height
            else:
                ws.row_dimensions[key].height = height
        logger.debug("Set row heights")

        for row in range(
            self.dimensions.border_rows[0], self.dimensions.border_rows[1] + 1
        ):
            for col in range(
                self.dimensions.border_cols[0], self.dimensions.border_cols[1] + 1
            ):
                ws.cell(row=row, column=col).border = self.styles.thin_border
        logger.debug(
            "Applied borders to rows %d-%d",
            self.dimensions.border_rows[0],
            self.dimensions.border_rows[1],
        )

    def initialize_headings(self) -> None:
        """Writes the company, contract, and week-of headings to the worksheet."""
        logger.debug("Setting headings section")
        if self.worksheet is None:
            raise RuntimeError("Worksheet not initialized")
        assert self.week_start is not None
        ws = self.worksheet
        ws["A1"] = "WSP USA, INC."
        ws["A1"].font = self.styles.title_font
        ws["A2"] = "NYSDOT 2025 Region 8 Bridge Inspection"
        ws["A2"].font = self.styles.title_font
        ws["A3"] = "Contract No. D037877"
        ws["A3"].font = self.styles.title_font

        ws["H1"] = "Inspection Schedule for Week of:"
        ws["H1"].font = self.styles.title_font
        ws.merge_cells("H1:I1")
        ws["H1"].alignment = Alignment(horizontal="right")
        ws["J1"] = self.week_start
        ws["J1"].font = self.styles.title_font
        ws["J1"].number_format = "mm/dd/yy"
        ws["J1"].alignment = Alignment(horizontal="center")
        ws["K1"] = self.week_start + timedelta(days=6)
        ws["K1"].font = self.styles.title_font
        ws["K1"].number_format = "mm/dd/yy"
        ws["K1"].alignment = Alignment(horizontal="center")
        ws["J2"] = "(Sunday)"
        ws["J2"].font = self.styles.normal_font
        ws["J2"].alignment = Alignment(horizontal="center")
        ws["K2"] = "(Saturday)"
        ws["K2"].font = self.styles.normal_font
        ws["K2"].alignment = Alignment(horizontal="center")

        ws.merge_cells("G3:H3")

    def save(self, filename: str | None = None) -> Path:
        """Saves the workbook to the output directory and returns the file path."""
        if filename is None:
            filename = self.default_filename
        output_path = self.output_dir / filename
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
        except PermissionError as e:
            logger.error(f"Permission denied creating directory: {e}")
            raise
        except OSError as e:
            logger.error(f"Failed to create directory: {e}")
            raise

        try:
            self.workbook.save(output_path)
            logger.info(f"Workbook saved: {output_path}")
        except PermissionError as e:
            logger.error(f"Permission denied saving file: {e}")
            raise
        except OSError as e:
            logger.error(f"Failed to save workbook: {e}")
            raise
        return output_path

    def to_bytes(self) -> bytes:
        """Returns the workbook content as bytes without writing to disk."""
        buf = BytesIO()
        self.workbook.save(buf)
        return buf.getvalue()


@dataclass
class WeeklyScheduleCreator(BaseCreator):
    """Creates a weekly inspection schedule pre-populated with bridge inspection data."""

    inspection_entries: list[InspectionEntry] = field(default_factory=list)

    def initialize_workbook(self) -> None:
        """Initializes workbook template then populates bridge data rows."""
        super().initialize_workbook()
        self.add_inspection_entries()

    def add_inspection_entries(self) -> None:
        """Populates table rows with bridge inspection data starting at row 25."""
        if self.worksheet is None:
            raise RuntimeError("Worksheet not initialized")
        ws = self.worksheet
        font = self.styles.table_font
        center = Alignment(horizontal="center", vertical="center")
        for i, entry in enumerate(self.inspection_entries):
            row = 25 + i
            values = [
                entry.team,
                entry.scheduled_date,
                entry.due_date,
                entry.region,
                entry.county,
                entry.bin,
                entry.feature_carried,
                entry.feature_crossed,
                entry.access,
                entry.town,
                entry.lane_closed,
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = font
                cell.alignment = center
                cell.border = self.styles.thin_border
                if isinstance(value, datetime):
                    cell.number_format = "mm/dd/yy"
        logger.debug("Populated %d data rows", len(self.inspection_entries))


def create_schedules_from_entries(
    entries: list[dict], team_name: str, output_dir: Path | None = None
) -> list[Path]:
    """Creates one weekly schedule file per week from parsed inspection entry dicts.

    All entries are attributed to team_name regardless of the source data's team column.
    Skips entries that fail InspectionEntry validation. Returns list of saved file paths.
    If output_dir is provided it overrides the default output/ directory.
    """
    valid: list[InspectionEntry] = []
    for e in entries:
        try:
            valid.append(InspectionEntry(**{**e, "team": team_name}))
        except Exception:
            logger.warning("Skipping invalid entry: %s", e.get("bin", "?"))

    weeks: dict[datetime, list[InspectionEntry]] = defaultdict(list)
    for entry in valid:
        week_start = BaseCreator.get_sunday(entry.scheduled_date)
        weeks[week_start].append(entry)

    paths = []
    for week_start, group in weeks.items():
        creator = WeeklyScheduleCreator(
            inspection_entries=group,
            week_start=week_start,
            team_name=team_name,
            output_dir_override=output_dir,
        )
        paths.append(creator.save())
        logger.info("Created schedule for team=%s week=%s: %d entries", team_name, week_start.date(), len(group))
    return paths


def create_schedules_as_bytes(
    entries: list[dict], team_name: str, output_dir: Path | None = None
) -> list[tuple[str, bytes]]:
    """Creates weekly schedule workbooks in memory, returns (filename, bytes) pairs.

    If output_dir is provided, also saves each file to disk at that path.
    Skips entries that fail InspectionEntry validation.
    """
    valid: list[InspectionEntry] = []
    for e in entries:
        try:
            valid.append(InspectionEntry(**{**e, "team": team_name}))
        except Exception:
            logger.warning("Skipping invalid entry: %s", e.get("bin", "?"))

    weeks: dict[datetime, list[InspectionEntry]] = defaultdict(list)
    for entry in valid:
        week_start = BaseCreator.get_sunday(entry.scheduled_date)
        weeks[week_start].append(entry)

    results = []
    for week_start, group in weeks.items():
        creator = WeeklyScheduleCreator(
            inspection_entries=group,
            week_start=week_start,
            team_name=team_name,
            output_dir_override=output_dir,
        )
        if output_dir:
            creator.save()
        results.append((creator.default_filename, creator.to_bytes()))
        logger.info("Created schedule for team=%s week=%s: %d entries", team_name, week_start.date(), len(group))
    return results


def create_sample_workbook() -> Path:
    """Creates an empty sample schedule workbook using DB-backed contact and team data."""
    creator = BaseCreator()
    return creator.save("sample_schedule.xlsx")
