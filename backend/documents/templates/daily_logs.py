from collections import defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from logging import getLogger
from pathlib import Path

from openpyxl import load_workbook

from documents.schemas import InspectionEntry
from documents.templates.base_doc import BaseCreator
from src.utils import get_sunday

logger = getLogger(__name__)


@dataclass(frozen=True)
class CellMapping:
    """Maps cell locations and inspection entry columns for a specific region."""

    region: str
    cells: dict[str, str] = field(default_factory=dict)
    entry_columns: dict[str, int] = field(default_factory=dict)

    def __post_init__(self):
        if self.region == "8":
            object.__setattr__(
                self,
                "cells",
                {
                    "team": "E1",
                    "date": "L1",
                    "license": "J16",
                    "odometer_reading": "L16",
                    "remarks": "C12",
                    "special_equipment": "A24",
                },
            )
            object.__setattr__(
                self,
                "entry_columns",
                {
                    "region": 1,
                    "county": 2,
                    "bin": 3,
                    "feature_carried": 4,
                    "feature_crossed": 5,
                    "arrival_time": 6,
                    "depart_time": 7,
                    "gen_rec": 8,
                    "posted_load": 9,
                    "flag": 10,
                    "wztc": 11,
                    "spec_emph": 12,
                    "access": 13,
                    "insp_type": 14,
                    "insp_comp": 15,
                },
            )


@dataclass
class DailyLogCreator(BaseCreator):
    """Loads daily log template and populates sheet-specific and common values."""

    team: str = "Chen"
    region: str = "8"
    cell_mapping: CellMapping = field(default_factory=lambda: CellMapping(region="8"))

    def __post_init__(self):
        self._init_directories()
        self.load_and_populate_template()

    def load_and_populate_template(self) -> None:
        """Loads template and populates team, dates, and region for Monday-Friday sheets."""
        template_path = (
            Path(__file__).resolve().parent.parent
            / "data"
            / f"R{self.region}_daily_log_template.xlsx"
        )
        logger.info(f"Loading template from {template_path}")
        self.workbook = load_workbook(template_path, keep_vba=True)

        sunday = get_sunday()
        monday = sunday + timedelta(days=1)
        days = ["MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY"]

        for i, day in enumerate(days):
            if day in self.workbook.sheetnames:
                current_date = monday + timedelta(days=i)
                self.populate_sheet_values(day, {"date": current_date})
                logger.debug(f"Populated {day} sheet")

        logger.info("Template populated successfully")

    def populate_sheet_values(
        self,
        sheet_name: str,
        day_values: dict[str, str | int | datetime] | dict[str, dict] | None = None,
    ) -> None:
        """Populates header and entry values for a sheet.

        Args:
            sheet_name: Name of the sheet to populate.
            day_values: Dict with header values or nested dict with 'header' and 'entries' keys.
                Simple format: {"date": datetime(...), "remarks": "text"}
                Nested format: {
                    "header": {"date": datetime(...), "remarks": "text"},
                    "entries": [{"region": "8", "county": "...", ...}, ...]
                }
        """
        if sheet_name not in self.workbook.sheetnames:
            logger.warning(f"Sheet {sheet_name} not found")
            return

        ws = self.workbook[sheet_name]
        mapping = self.cell_mapping

        ws[mapping.cells["team"]] = self.team
        logger.debug(f"Set team '{self.team}' at {mapping.cells['team']}")

        if not day_values:
            return

        if "header" in day_values and "entries" in day_values:
            self._populate_header(ws, mapping, day_values.get("header", {}))  # type: ignore
            self._populate_entries(ws, mapping, day_values.get("entries", []))  # type: ignore
        else:
            self._populate_header(ws, mapping, day_values)  # type: ignore

    def _populate_header(
        self, ws, mapping: CellMapping, header_values: dict
    ) -> None:
        """Populates header cells (team, date, remarks, etc.) for a sheet."""
        for field, value in header_values.items():
            if field in mapping.cells:
                ws[mapping.cells[field]] = value
                logger.debug(f"Set {field} '{value}' at {mapping.cells[field]}")
            else:
                logger.debug(f"Field '{field}' not in cell mapping")

    def _populate_entries(self, ws, mapping: CellMapping, entries: list) -> None:
        """Populates inspection entry rows (4-10) for a sheet."""
        start_row = 4
        max_rows = 7

        for i, entry in enumerate(entries):
            if i >= max_rows:
                logger.warning(f"Sheet has max {max_rows} entry rows; skipping entry {i}")
                break

            row = start_row + i
            for field, col_num in mapping.entry_columns.items():
                if field in entry:
                    value = entry[field]
                    cell = ws.cell(row=row, column=col_num, value=value)
                    if isinstance(value, datetime):
                        cell.number_format = "mm/dd/yy"
                    logger.debug(f"Set {field} '{value}' at row {row}, col {col_num}")

        logger.debug(f"Populated {min(len(entries), max_rows)} entry rows")

    def populate_weekly_log(self, schedule_data: list[dict]) -> None:
        """Populates daily log sheets with inspection entries grouped by scheduled date.

        Args:
            schedule_data: List of inspection entry dicts from weekly_schedule.py
                containing fields like region, county, bin, feature_carried,
                feature_crossed, scheduled_date, etc.
        """
        valid_entries: list[InspectionEntry] = []
        for entry in schedule_data:
            try:
                valid_entries.append(InspectionEntry(**entry))
            except Exception as e:
                logger.warning(f"Skipping invalid entry {entry.get('bin', '?')}: {e}")

        # Group entries by scheduled date
        entries_by_date: dict[datetime, list[dict]] = defaultdict(list)
        for entry in valid_entries:
            date_key = entry.scheduled_date.date()
            entries_by_date[date_key].append(entry.model_dump())

        # Map dates to sheet names
        day_names = ["SUNDAY", "MONDAY", "TUESDAY", "WEDNESDAY", "THURSDAY", "FRIDAY", "SATURDAY"]

        for date_key, entries in sorted(entries_by_date.items()):
            day_of_week = date_key.weekday()
            sheet_name = day_names[(day_of_week + 1) % 7]  # Convert Python weekday to sheet day

            if sheet_name in self.workbook.sheetnames:
                day_values = {
                    "header": {"date": datetime.combine(date_key, datetime.min.time())},
                    "entries": entries,
                }
                self.populate_sheet_values(sheet_name, day_values)
                logger.info(f"Populated {sheet_name} with {len(entries)} entries for {date_key}")
            else:
                logger.warning(f"Sheet {sheet_name} not found for date {date_key}")


def create_daily_logs_from_schedule(
    schedule_data: list[dict], team_name: str, output_dir: Path | None = None
) -> list[Path]:
    """Creates weekly daily logs with inspection entries prefilled from schedule data.

    Args:
        schedule_data: List of inspection entry dicts from weekly_schedule.py.
        team_name: Team name to populate in the daily logs.
        output_dir: Optional output directory; uses default if None.

    Returns:
        List of saved file paths for each week's daily log.
    """
    weeks: dict[datetime, list[dict]] = defaultdict(list)
    for entry in schedule_data:
        try:
            inspection_entry = InspectionEntry(**entry)
            week_start = get_sunday(inspection_entry.scheduled_date)
            weeks[week_start].append(entry)
        except Exception as e:
            logger.warning(f"Skipping invalid entry {entry.get('bin', '?')}: {e}")

    paths = []
    for week_start, entries in sorted(weeks.items()):
        creator = DailyLogCreator(team=team_name, output_dir_override=output_dir)
        creator.populate_weekly_log(entries)
        paths.append(creator.save())
        logger.info(
            f"Created daily logs for team={team_name} week={week_start.date()} "
            f"with {len(entries)} total entries"
        )
    return paths


def create_daily_logs_as_bytes(
    entries: list[dict], team_name: str
) -> list[tuple[str, bytes]]:
    """Creates weekly daily log workbooks in memory, returns (filename, bytes) pairs.

    Skips entries that fail InspectionEntry validation.
    """
    weeks: dict[datetime, list[dict]] = defaultdict(list)
    for entry in entries:
        try:
            inspection_entry = InspectionEntry(**entry)
            week_start = get_sunday(inspection_entry.scheduled_date)
            weeks[week_start].append(entry)
        except Exception:
            logger.warning(f"Skipping invalid entry: {entry.get('bin', '?')}")

    results = []
    for week_start, week_entries in sorted(weeks.items()):
        creator = DailyLogCreator(team=team_name)
        creator.populate_weekly_log(week_entries)
        week_str = week_start.strftime("%-m-%-d-%y")
        filename = f"{team_name} Region 8 Daily Logs - Week of {week_str}.xlsm"
        results.append((filename, creator.to_bytes()))
        logger.info(
            f"Created daily logs for team={team_name} week={week_start.date()} "
            f"with {len(week_entries)} total entries"
        )
    return results
